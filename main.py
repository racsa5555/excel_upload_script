import os

import pandas as pd

from dotenv import load_dotenv

import gspread
from google.oauth2.service_account import Credentials
from gspread.exceptions import APIError

from openpyxl import load_workbook

load_dotenv()

scopes = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

KEY_FILE = 'credentials.json'

#RGB #7FFF00
COLOR = {
        "red": 0.498,
        "green": 1.0,
        "blue": 0.0
}

GOOGLE_SHEET_NAME = os.getenv('GOOGLE_SHEET_NAME')
folder_path = os.getenv('PATH_TO_FOLDER')

credentials = Credentials.from_service_account_file(KEY_FILE, scopes=scopes)

client = gspread.authorize(credentials)


files = os.listdir(folder_path)

excel = files[0]

excel_name = files[0].split('.')[0]

spreadsheet = client.open(title = GOOGLE_SHEET_NAME)


xls = pd.ExcelFile(folder_path+'/'+ excel)


#Итерация по листам
for sheet_name in xls.sheet_names:
    df = pd.read_excel(xls, sheet_name=sheet_name)

    worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=df.shape[0], cols=df.shape[1])

    #Чтение данных с листа
    data = [df.columns.values.tolist()] + df.values.tolist()

    #Очистка nan значений
    cleaned_data = [[cell if pd.notna(cell) else '' for cell in row] for row in data]  

    worksheet.update(values=cleaned_data)

    wb = load_workbook(folder_path+'/'+ excel)
    
    sheet = wb[sheet_name]

    lines_with_color = []
    
    # Поиск координатов ячеек с цветом #7FFF00
    for row in sheet.iter_rows():
        line_with_color = []
        for cell in row:
            if cell.fill.start_color.rgb == 'FF7FFF00':
                line_with_color.append(cell.coordinate)
        lines_with_color.append(line_with_color)
        line_with_color = []

    lines_with_color = list(filter(lambda x: len(x) != 0,lines_with_color))
    
    # Обновление цвета в google_sheets
    for cell in lines_with_color:
        worksheet.format(f'A{str(cell[0][-1])}:{cell[-1]}',{"backgroundColor":COLOR})


print('Данные успешно перенесены из Excel в Google таблицу')
